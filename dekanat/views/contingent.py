# -*- coding: utf-8 -*-
import json
import tempfile
from datetime import datetime, timedelta

import openpyxl
from Scripts._testcapi import return_result_with_error
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.shortcuts import render
from django.template.loader import render_to_string
from django.utils import dateformat
from docxtpl import DocxTemplate
from openpyxl.writer.excel import save_virtual_workbook

from dekanat.controllers.contingent_movement_report import create_movement_xlsx_report
from dekanat.controllers.dismissed_and_transfered_report import create_dismissed_and_transfered_report
from dekanat.controllers.serialize_student import serialize_student
from dekanat.controllers.student_reference import get_reference_data, get_bypass_data, get_military_students_data, \
    get_study_edu_year_cert_data
from dekanat.controllers.students_filter import get_students_fields
from dekanat.controllers.tree_utils import get_tree
from dekanat.controllers.views_controllers import ViewDepartment, ViewSpec, ViewGroup
from dekanat.controllers.vpo_reports import get_vpo_report_2_1_2, get_vpo_report_2_13, get_vpo_report_2_1_5, \
    get_vpo_foreign_by_spec
from dekanat.controllers.vpo_reports.report_2_12 import get_vpo_report_2_12
from dekanat.models.documents import Reference, StudyPeriodCertificate
from dekanat.models.group import Group
from dekanat.models.student import Student
from dekanat.views.utils import create_xlsx_response
from dict_app.models_.depart import Department, EduDepartment
from dict_app.models_.depart_settings import UniversitySettings
from dict_app.models_.dict import DictEduType
from edu_plans.models.spec import Spec
from history_app.controllers.student_controller import get_student_state
from orders.models import PointOrderInAcadem
from print.file_response_utils import get_file_response
from print.print_reports.age_ref import AgeRef
from print.print_reports.bypass_inst import BypassInst
from print.print_reports.bypass_izo import BypassIzo
from print.print_reports.bypass_printer import BypassPrinter
from print.print_reports.military_ref_asp import MilitaryRefAsp
from print.print_reports.military_ref_student import MilitaryRefStudent
from print.print_reports.military_ref_vk import MilitaryRefVK


def get_filter_query(filter_info, value):
    filter_kwargs = dict()
    if filter_info['filter_type'] == 'select':
        if filter_info['json_field'] == 'mil_office' and value[0] == '-1':
            return ~Q(**{filter_info['db_field'] + '__in': [38, 37, 51, 50]}) & \
                   Q(**{filter_info['db_field'] + '__isnull': False})
        if filter_info['json_field'] == 'is_snils':
            if value[0] == '1':
                return Q(**{filter_info['db_field'] + '__isnull': False}) & ~Q(**{filter_info['db_field']: ''})
            elif value[0] == '0':
                return Q(**{filter_info['db_field'] + '__isnull': True}) | Q(
                    **{filter_info['db_field'] + '__exact': ''})
        if len(value) > 1:
            filter_kwargs[filter_info['db_field'] + '__in'] = value
        else:
            filter_kwargs[filter_info['db_field']] = value[0]
    elif filter_info['filter_type'] == 'bool':
        if filter_info['json_field'] in ('is_starosta', 'debt', 'is_snils'):
            filter_kwargs[filter_info['db_field']] = False
        else:
            filter_kwargs[filter_info['db_field']] = True
    elif filter_info['filter_type'] == 'date':
        period = value.split(' - ')
        if len(period) > 1:
            date_start = datetime.strptime(period[0], '%d.%m.%Y')
            date_end = datetime.strptime(period[1], '%d.%m.%Y')
            filter_kwargs[filter_info['db_field'] + '__gte'] = date_start
            filter_kwargs[filter_info['db_field'] + '__lte'] = date_end
        else:
            date = datetime.strptime(period[0], '%d.%m.%Y')
            filter_kwargs[filter_info['db_field']] = date
    elif filter_info['json_field'] == 'fact_address' or filter_info['json_field'] == 'reg_address':
        return Q() | \
               Q(**{filter_info['db_field'] + '__region': value}) | \
               Q(**{filter_info['db_field'] + '__district': value}) | \
               Q(**{filter_info['db_field'] + '__settlement': value}) | \
               Q(**{filter_info['db_field'] + '__street': value}) | \
               Q(**{filter_info['db_field'] + '__district': value})
    else:
        filter_kwargs[filter_info['db_field'] + '__icontains'] = value

    return Q(**filter_kwargs)


def get_students(dep_id, node_id, type=0, status=1, filter_fields=None, filter_params=()):
    q = Q()
    for param in filter_params:
        filter_info = next((item for item in filter_fields if item['json_field'] == param['name']), False)
        if not filter_info:
            return HttpResponseBadRequest('Имя фильтра {} не найдено'.format(param['name']))
        q = q & get_filter_query(filter_info, param['value'])
    institute = Department.objects.get(id=dep_id)
    if type == 0 or type == 1:  # "Department" или кафедра:
        dep = Department.objects.get(id=node_id)
        return ViewDepartment(dep).get_all_student(status, q)
    if type == 2:  # "Spec":
        spec = Spec.objects.get(id=node_id)
        return ViewSpec(spec, institute).get_all_student(status, q)
    if type == 3:  # "Group":
        group = Group.objects.get(id=node_id)
        return ViewGroup(group).get_all_student(status, q)
    if type == 8:  # "EduForm":
        return ViewDepartment(institute).get_all_student(status, q).filter(plan__spec__eduForm=node_id)


def get_students_table(dep_id, node_id, type, status, page=1, filter_fields=None, filter_params=()):
    students = get_students(dep_id, int(node_id), int(type), int(status), filter_fields, filter_params)
    students_pagination = Paginator(students, 30).get_page(page)
    students_dict = []
    fields = [field['json_field'] for field in get_students_fields() if 'json_field' in field]
    for student in students_pagination:
        students_dict.append(serialize_student(student, fields))
    return {'students': json.dumps(students_dict),
            'pagination': render_to_string('contingent/contingent_pagination.html',
                                           dict(pagination=students_pagination))}


def get_students_table_response(request, dep_id, node_id, type, status):
    filter_params = json.loads(request.GET.get('filter', '[]'))
    page = int(request.GET.get('page', '1'))
    filter_fields = get_students_fields(request.curr_dep)
    return JsonResponse(get_students_table(dep_id, node_id, type, status, page, filter_fields, filter_params))


def open_all_contingent(request):
    dep = request.curr_dep
    view_type = request.GET.get('viewtype', None)
    tree = get_tree('contingent', dep, view_type=view_type)
    students_table = get_students_table(dep.id, dep.id, 0, 1, 1)
    edu_types = DictEduType.objects.all()
    return render(request, 'contingent/contingent.html',
                  dict(tree=tree, students_table=students_table, view_type=view_type, edu_types=edu_types,
                       fields=json.dumps(get_students_fields(dep))))


def get_students_by_date(dep_id, date, node_id, type=0, status=1, filter_fields=None, filter_params=()):
    q_hist = Q()
    q_no_hist = Q()
    for param in filter_params:
        filter_info = next((item for item in filter_fields if item['json_field'] == param['name']), False)
        if not filter_info:
            return HttpResponseBadRequest('Имя фильтра {} не найдено'.format(param['name']))
        q = get_filter_query(filter_info, param['value'])
        if filter_info.get('history', False):
            q_hist = q_hist & q
        else:
            q_no_hist = q_no_hist & q

    students = []
    institute = Department.objects.get(id=dep_id)
    if type == 0:  # "Department":
        dep = Department.objects.get(id=node_id)
        students.extend(ViewDepartment(dep).get_all_students_in_history(date, status, q_no_hist, q_hist))
    if type == 1:  # Kafedra
        kaff = Department.objects.get(id=node_id)
        students.extend(ViewDepartment(kaff).get_all_students_in_history(date, status, q_no_hist, q_hist))
    if type == 2:  # "Spec":
        spec = Spec.objects.get(id=node_id)
        students.extend(ViewSpec(spec, institute).get_all_students_in_history(date, status, q_no_hist, q_hist))
    if type == 3:  # "Group":
        group = Group.objects.get(id=node_id)
        students.extend(ViewGroup(group).get_all_students_in_history(date, status, q_no_hist, q_hist))
    if type == 8:  # "EduForm":
        q_hist = q_hist & Q(plan__spec__eduForm=node_id)
        return ViewDepartment(institute).get_all_students_in_history(date, status, q_no_hist, q_hist)
    return students


def get_students_table_by_date(dep_id, date, id, type, status, page, filter_fields=None, request_params=()):
    students = get_students_by_date(dep_id, date, int(id), int(type), int(status), filter_fields, request_params)
    students_pagination = Paginator(students, 30).get_page(page)
    students_dict = []
    fields = [field['json_field'] for field in get_students_fields() if 'json_field' in field]
    for student in students_pagination:
        student = get_student_state(student.student, student)
        students_dict.append(serialize_student(student, fields))

    return {'students': json.dumps(students_dict),
            'pagination': render_to_string('contingent/contingent_pagination.html',
                                           dict(pagination=students_pagination))}


def get_students_table_by_date_response(request, date, dep_id, node_id, type, status):
    date = datetime.strptime(date, '%d.%m.%Y').date()
    request_params = json.loads(request.GET.get('filter', '[]'))
    page = int(request.GET.get('page', '1'))
    filter_fields = get_students_fields(request.curr_dep)  # todo доделать для университет
    return JsonResponse(
        get_students_table_by_date(dep_id, date, node_id, type, status, page, filter_fields, request_params))


def open_all_contingent_by_date(request, date):
    date = datetime.strptime(date, '%d.%m.%Y').date()
    dep = request.curr_dep
    view_type = request.GET.get('viewtype', None)
    tree = get_tree('contingent', dep, date, view_type)
    students_table = get_students_table_by_date(dep.id, date, dep.id, type=0, status=1, page=1)
    return render(request, 'contingent/contingent.html',
                  dict(tree=tree, students_table=students_table, date=date, view_type=view_type,
                       fields=json.dumps(get_students_fields(dep))))


def get_request_students(request, dep_id, node_id, type, status, date):
    students = request.GET.getlist('students', [])

    if students and students[0]:
        return Student.objects.filter(id__in=students)

    request_params = json.loads(request.GET.get('filter', '[]'))
    filter_fields = get_students_fields(request.curr_dep)
    if date:
        date = datetime.strptime(date, '%d.%m.%Y').date()
        students = []
        hist_students = get_students_by_date(dep_id, date, int(node_id), int(type), int(status),
                                             filter_fields, request_params)
        for student in hist_students:
            student = get_student_state(student.student, student)
            students.append(student)
        return students
    return get_students(dep_id, int(node_id), int(type), int(status), filter_fields, request_params)


def _create_docx_response(context, template_name, result_file_name):
    doc = DocxTemplate(template_name)
    doc.render(context, autoescape=True)
    temp = tempfile.NamedTemporaryFile(dir='tmp', delete=False)
    doc.save(temp.name)
    f = open(temp.name, 'rb').read()
    response = HttpResponse(f)
    response['mimetype'] = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    response['Content-Disposition'] = f'attachment; filename={result_file_name}.docx'
    return response


def _get_print_reference_context(request):
    student_id = int(request.POST.get('student'))
    student = Student.objects.get(id=student_id)
    options = dict(is_study_period='study_period' in request.POST,
                   is_orders='orders_info' in request.POST,
                   is_spec='spec_info' in request.POST,
                   is_director='sign_director' in request.POST,
                   is_deputy_director='sign_deputy_director' in request.POST,
                   is_edu_vicerector='sign_edu_vicerector' in request.POST,
                   is_first_vicerector='sign_first_vicerector' in request.POST,
                   targer_place=request.POST.get('targer_place'),
                   is_dotted_line='dotted_line' in request.POST,
                   response_number=request.POST.get('response_number'))

    if request.POST.get('response_date'):
        options['response_date'] = datetime.strptime(request.POST.get('response_date'), '%d.%m.%Y')
    student_data = get_reference_data(student, options)
    last_ref_number = student.group.dep.get_field_settings('last_reference_number') + 1
    ref_number = '{}/{}{}'.format(str(datetime.now().year)[2:], student.group.dep.get_field_settings('code'),
                                  last_ref_number)

    if "is_register" in request.POST:
        student.group.dep.set_field_settings('last_reference_number', last_ref_number)
        Reference.objects.create(student=student, place=options['targer_place'], number=ref_number)

    university_settings = student.group.dep.get_university().universitysettings

    return dict(student_data=student_data, options=options, ref_number=ref_number,
                university_settings=university_settings)


def print_references(request):
    context = _get_print_reference_context(request)
    return render(request, f'student_reference/{context["student_data"]["template_name"]}.html', context)


def print_references_docx(request):
    context = _get_print_reference_context(request)
    return _create_docx_response(context, "template_report/student_ref_template.docx", 'student_ref')


def log_register_reference(request):
    """
    Таблица журнала выданных справок об обучении
    """
    dep = request.curr_dep
    students = get_students(request.curr_dep, int(dep.id))
    references = Reference.objects.filter(student__in=students).order_by('-date', '-number')
    return render(request, f'log_register_reference.html', dict(references=references))


def print_study_period_certificate(request, id=1, type=0, status=1, date=None):
    """
    Печать справки о периоде обучения
    """
    student_id = int(request.POST.get('student'))
    student = Student.objects.get(id=student_id)
    params = get_study_edu_year_cert_data(student)
    params['continue'] = request.POST.get('student_ref_status') == 'continue'
    params['required'] = 'is_place' in request.POST

    if "is_register" in request.POST:
        ref_number = request.POST.get('reg_number')
        ref_date = datetime.strptime(request.POST.get('issue_date'), '%d.%m.%Y')
        params['number'] = ref_number
        params['date'] = dateformat.format(ref_date, 'j E Y')
        StudyPeriodCertificate.objects.create(student=student, number=ref_number, date=ref_date)

    ar = AgeRef(str(datetime.now().timestamp()) + ' справка о периоде обучения.pdf', params)
    ar.print()
    return get_file_response(request, ar.file_name, False)


def print_bypasses(request, dep_id, node_id=1, type=0, status=1, date=None):
    data_list = []
    students = get_request_students(request, dep_id, node_id, type, status, date)
    for student in students:
        data_list.append(get_bypass_data(student))

    f_student = students.first()
    if f_student.plan.spec.spec.level.level == 6:
        bp = BypassPrinter(str(datetime.now().timestamp()) + ' обходные.pdf', data_list)
    elif f_student.plan.spec.eduForm.edu_form == 1:
        bp = BypassInst(str(datetime.now().timestamp()) + ' обходные.pdf', data_list)
    else:
        bp = BypassIzo(str(datetime.now().timestamp()) + ' обходные.pdf', data_list)

    bp.print()
    return get_file_response(request, bp.file_name, False)


def export_excel(request, dep_id, node_id=1, type=0, status=1, date=None):
    students = get_request_students(request, dep_id, node_id, type, status, date)
    columns = request.GET.getlist('columns')
    wb = openpyxl.Workbook()
    sheet = wb.active
    column_start = 1
    if 'fio_concat' in request.GET:
        sheet.cell(row=1, column=column_start, value='ФИО')
        column_start = 2

    for i, column in enumerate(columns, start=column_start):
        label = next((item['label'] for item in get_students_fields() if item['json_field'] == column), False)
        sheet.cell(row=1, column=i, value=label)

    for i, student in enumerate(students, start=2):
        student_data = serialize_student(student, columns)
        if 'fio_concat' in request.GET:
            sheet.cell(row=i, column=1, value=str(student))
        for j, column in enumerate(columns, start=column_start):
            sheet.cell(row=i, column=j, value=student_data[column])

    response = HttpResponse(content=save_virtual_workbook(wb))
    response['mimetype'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response['Content-Disposition'] = 'attachment; filename=contingent.xlsx'
    return response


def _get_print_military_context(request, dep_id, node_id=1, type=0, status=1, date=None, mil_ref_date=None):
    if not mil_ref_date:
        mil_ref_date = datetime.now()
    students = get_request_students(request, dep_id, node_id, type, status, date)
    students_data = get_military_students_data(students)
    if hasattr(request.curr_dep, 'university'):
        bstu_settings = request.curr_dep.university.universitysettings
    else:
        bstu_settings = request.curr_dep.edudepartment.get_institute().dep.university.universitysettings
    general_data = {
        'date': dateformat.format(mil_ref_date, '«j» E Y г.'),
        'acc_number': bstu_settings.license_number,
        'acc_date_start': dateformat.format(bstu_settings.license_date, 'd.m.Y г.'),
        'acc_date_finish': dateformat.format(datetime(bstu_settings.license_date.year + 6,
                                                      bstu_settings.license_date.month,
                                                      bstu_settings.license_date.day), 'd.m.Y г.'),
        'vicerector': bstu_settings.vicerector,
        'vicerector_science': bstu_settings.vicerector_science
    }
    return {'students_data': students_data, 'general_data': general_data}


def print_military_ref(request, dep_id, node_id=1, type=0, status=1):
    mil_ref_date = datetime.strptime(request.GET.get('mil_ref_date'), '%d.%m.%Y')
    context = _get_print_military_context(request, dep_id, node_id, type, status, None, mil_ref_date)
    data = []
    for student_data in context['students_data']:
        student_data = {**student_data, **context['general_data']}
        if 'mil_ref_spec' in request.GET:
            student_data['spec'] = request.GET.get('mil_ref_spec')
        if 'mil_ref_kurs' in request.GET:
            student_data['kurs'] = request.GET.get('mil_ref_kurs')
        data.append(student_data)

    if data[0]['is_vk']:
        ref_printer = MilitaryRefVK(str(datetime.now().timestamp()) + ' справки для военкомата.pdf', data)
    else:
        if data[0]['level'] == 6:
            ref_printer = MilitaryRefAsp(str(datetime.now().timestamp()) + ' справки для военкомата.pdf', data)
        else:
            ref_printer = MilitaryRefStudent(str(datetime.now().timestamp()) + ' справки для военкомата.pdf', data)
    ref_printer.print()

    return get_file_response(request, ref_printer.file_name, False)


def get_mil_ref_docx(request, dep_id, node_id=1, type=0, status=1, date=None):
    context = _get_print_military_context(request, dep_id, node_id, type, status, date)
    return _create_docx_response(context, 'template_report/mil_ref_template.docx', 'mil_refs')


def get_movement_report(request):
    dep = request.GET.get('dep')
    dep = EduDepartment.objects.get(id=dep)
    period = request.GET.get('report_period').split(' - ')
    start_date = datetime.strptime(period[0], '%d.%m.%Y').date()
    end_date = datetime.strptime(period[1], '%d.%m.%Y').date()
    file_name = create_movement_xlsx_report(dep, start_date, end_date)
    return create_xlsx_response(file_name, 'movement_report')


def get_dismissed_transfered_report(request):
    dep = request.GET.get('dep')
    dep = Department.objects.get(id=dep)
    if hasattr(dep, 'university'):
        institutes_id = 2, 3, 4, 6, 7, 10, 13
    else:
        institutes_id = dep.id,
    period = request.GET.get('report_period').split(' - ')
    start_date = datetime.strptime(period[0], '%d.%m.%Y').date()
    end_date = datetime.strptime(period[1], '%d.%m.%Y').date()
    file_name = create_dismissed_and_transfered_report(start_date, end_date, institutes_id)
    result_file_name = f'R041001061697_{datetime.now().month}{str(datetime.now().year)[2:]}'
    return create_xlsx_response(file_name, result_file_name)


def vpo_report_2_1_2(request, edu_form):
    file_name = get_vpo_report_2_1_2(edu_form)
    result_file_name = f'vpo_2_1_2_{"full_time" if edu_form == "1" else "part_time"}'
    return create_xlsx_response(file_name, result_file_name)


def vpo_report_2_1_5(request, edu_form):
    file_name = get_vpo_report_2_1_5(edu_form)
    result_file_name = f'vpo_2_1_5_{"full_time" if edu_form == "1" else "part_time"}'
    return create_xlsx_response(file_name, result_file_name)


def vpo_report_2_12(request, edu_form):
    file_name = get_vpo_report_2_12(edu_form)
    result_file_name = f'vpo_2_12_{"full_time" if edu_form == "1" else "part_time"}'
    return create_xlsx_response(file_name, result_file_name)


def vpo_report_2_13(request, edu_form):
    file_name = get_vpo_report_2_13(edu_form)
    result_file_name = f'vpo_2_13_{"full_time" if edu_form == "1" else "part_time"}'
    return create_xlsx_response(file_name, result_file_name)


def vpo_report_foreign_by_spec(request):
    file_name = get_vpo_foreign_by_spec()
    return create_xlsx_response(file_name, 'vpo_foreign_by_spec')


def student_in_academ_report(request):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1, value="ФИО")
    sheet.cell(row=1, column=2, value="Группа")
    sheet.cell(row=1, column=3, value="№ зачётки")
    sheet.cell(row=1, column=4, value="Основа обучения")
    sheet.cell(row=1, column=5, value="Дата начала академ. отпуска")
    sheet.cell(row=1, column=6, value="Дата окончания академ. отпуска")
    sheet.cell(row=1, column=7, value="Причина ухода в академ. отпуск")

    deps = list(ViewDepartment(request.curr_dep).get_child_deps_edu()) + [request.curr_dep]
    point_history = PointOrderInAcadem.objects.filter(order__department__in=deps, student__status=4).order_by(
        'order__department', 'group', 'student')

    for i, history in enumerate(point_history, start=2):
        sheet.cell(row=i, column=1, value=str(history.student))
        sheet.cell(row=i, column=2, value=str(history.group))
        sheet.cell(row=i, column=3, value=str(history.student.zcode))
        sheet.cell(row=i, column=4, value=str(history.student.edu_type))
        sheet.cell(row=i, column=5, value=history.in_academ_date.strftime('%d.%m.%Y'))
        sheet.cell(row=i, column=6,
                   value='' if str(history.out_academ_date) == 'None' else history.out_academ_date.strftime('%d.%m.%Y'))
        sheet.cell(row=i, column=7, value=str('' if str(history.reason) == 'None' else history.reason))

    for col in sheet.columns:
        length = max(len(str(cell.value)) for cell in col)
        sheet.column_dimensions[col[0].column_letter].width = length

    response = HttpResponse(content=save_virtual_workbook(wb))
    response['mimetype'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response['Content-Disposition'] = 'attachment; filename=student_in_academ.xlsx'
    return response


def analitic(request):
    tree = get_tree('contingent', request.curr_dep, view_type=2)
    return render(request, 'analitics.html')
